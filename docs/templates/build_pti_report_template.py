#!/usr/bin/env python3
"""Generate PTI rescue report Excel template with formulas."""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/admin/Documents/RSA/rsa-design/docs/templates/pti-rescue-report-template.xlsx"

HEADERS = [
    "STT",
    "Biển kiểm soát",
    "Loại dịch vụ",
    "Trạng thái bảo lãnh (nguồn)",
    "PTI thanh toán",
    "KH đã thanh toán",
    "KH công nợ",
    "Tổng chi phí",
    "Nhóm báo cáo",
]

# Sample rows mirroring user's pivot (amounts in VND)
SAMPLE_ROWS = [
    ("77x", "Cẩu kéo", "PTI bảo lãnh", 128_370_878, 0, 0),
    ("4x", "Hỗ trợ tại chỗ", "PTI bảo lãnh", 1_600_000, 0, 0),
    ("1x", "Cẩu kéo", "PTI bảo lãnh 1 phần", 2_997_172, 788_889, 0),
    ("2x", "Cẩu kéo", "PTI không bảo lãnh", 0, 3_445_000, 0),
    ("1x", "Cẩu kéo, thủy kích", "PTI không bảo lãnh", 0, 1_573_000, 0),
    ("2x", "Đâm, lật, tai nạn", "PTI không bảo lãnh", 0, 25_359_259, 0),
    ("1x", "Hỗ trợ tại chỗ", "PTI không bảo lãnh", 0, 725_926, 0),
    ("1x", "Đâm, lật, tai nạn", "PTI không bảo lãnh, Khách hàng chưa thanh toán", 0, 0, 11_666_667),
]

GROUPS = [
    "Bảo lãnh đủ",
    "Bảo lãnh một phần",
    "Không bảo lãnh (KH đã TT)",
    "Không bảo lãnh (Chưa thu)",
]

SERVICES = [
    "Cẩu kéo",
    "Hỗ trợ tại chỗ",
    "Cẩu kéo, thủy kích",
    "Đâm, lật, tai nạn",
]

thin = Side(style="thin", color="B4C6E7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
header_fill = PatternFill("solid", fgColor="D9E1F2")
total_fill = PatternFill("solid", fgColor="E2EFDA")
title_font = Font(bold=True, size=12)
header_font = Font(bold=True)
money_fmt = "#,##0"
count_fmt = "#,##0"


def style_range(ws, cell_range, fill=None, font=None):
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border
            if fill:
                cell.fill = fill
            if font:
                cell.font = font


def build_data_sheet(wb):
    ws = wb.active
    ws.title = "DuLieu"
    ws["A1"] = "Dữ liệu chi tiết đơn cứu hộ — thay bằng export từ hệ thống"
    ws["A1"].font = title_font
    ws.merge_cells("A1:I1")

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    row = 4
    stt = 1
    for marker, service, status, pti, kh_paid, kh_debt in SAMPLE_ROWS:
        count = int(marker.replace("x", ""))
        per_case_pti = pti // count if count else pti
        per_case_kh = kh_paid // count if count else kh_paid
        per_case_debt = kh_debt // count if count else kh_debt
        for _ in range(count):
            ws.cell(row=row, column=1, value=stt)
            ws.cell(row=row, column=2, value=f"BIển-{stt:03d}")
            ws.cell(row=row, column=3, value=service)
            ws.cell(row=row, column=4, value=status)
            ws.cell(row=row, column=5, value=per_case_pti)
            ws.cell(row=row, column=6, value=per_case_kh)
            ws.cell(row=row, column=7, value=per_case_debt)
            ws.cell(row=row, column=8, value=f"=E{row}+F{row}+G{row}")
            ws.cell(row=row, column=9, value=(
                f'=IF(D{row}="PTI bảo lãnh","Bảo lãnh đủ",'
                f'IF(D{row}="PTI bảo lãnh 1 phần","Bảo lãnh một phần",'
                f'IF(D{row}="PTI không bảo lãnh","Không bảo lãnh (KH đã TT)",'
                f'IF(D{row}="PTI không bảo lãnh, Khách hàng chưa thanh toán",'
                f'"Không bảo lãnh (Chưa thu)",D{row}))))'
            ))
            for c in range(1, 10):
                ws.cell(row=row, column=c).border = border
            for c in (5, 6, 7, 8):
                ws.cell(row=row, column=c).number_format = money_fmt
            row += 1
            stt += 1

    last_row = row - 1
    ws.cell(row=row + 1, column=4, value="LAST_ROW marker:")
    ws.cell(row=row + 1, column=5, value=last_row)

    widths = [6, 16, 22, 42, 16, 16, 14, 16, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    return last_row


def build_report_sheet(wb, last_row):
    ws = wb.create_sheet("BaoCao")
    ws["A1"] = "BÁO CÁO CỨU HỘ PTI"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Kỳ báo cáo:"
    ws["B2"] = "01/07/2026 – 31/07/2026"
    ws["A3"] = "Đơn vị lập:"
    ws["B3"] = "VETC RSA"

    # --- Section 1: Summary by guarantee group ---
    ws["A5"] = "1. Tổng quan theo trạng thái bảo lãnh"
    ws["A5"].font = title_font

    summary_headers = [
        "Nhóm báo cáo",
        "Số ca",
        "Tổng chi phí",
        "PTI thanh toán",
        "KH đã thanh toán",
        "KH công nợ",
        "Ghi chú",
    ]
    for col, h in enumerate(summary_headers, start=1):
        cell = ws.cell(row=6, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    notes = {
        "Bảo lãnh đủ": "",
        "Bảo lãnh một phần": "PTI và KH cùng chi trả phần vượt hạn mức",
        "Không bảo lãnh (KH đã TT)": "KH tự chi trả toàn bộ",
        "Không bảo lãnh (Chưa thu)": "Đang theo dõi thu hồi",
    }

    r = 7
    for group in GROUPS:
        ws.cell(row=r, column=1, value=group).border = border
        ws.cell(row=r, column=2, value=f'=COUNTIF(DuLieu!$I$4:$I${last_row},A{r})').border = border
        ws.cell(row=r, column=3, value=f'=SUMIF(DuLieu!$I$4:$I${last_row},A{r},DuLieu!$H$4:$H${last_row})').border = border
        ws.cell(row=r, column=4, value=f'=SUMIF(DuLieu!$I$4:$I${last_row},A{r},DuLieu!$E$4:$E${last_row})').border = border
        ws.cell(row=r, column=5, value=f'=SUMIF(DuLieu!$I$4:$I${last_row},A{r},DuLieu!$F$4:$F${last_row})').border = border
        ws.cell(row=r, column=6, value=f'=SUMIF(DuLieu!$I$4:$I${last_row},A{r},DuLieu!$G$4:$G${last_row})').border = border
        ws.cell(row=r, column=7, value=notes[group]).border = border
        for c in (2, 3, 4, 5, 6):
            ws.cell(row=r, column=c).number_format = money_fmt if c != 2 else count_fmt
        r += 1

    ws.cell(row=r, column=1, value="TỔNG").font = header_font
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = total_fill
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).font = header_font
    ws.cell(row=r, column=2, value=f"=SUM(B7:B{r-1})")
    ws.cell(row=r, column=3, value=f"=SUM(C7:C{r-1})")
    ws.cell(row=r, column=4, value=f"=SUM(D7:D{r-1})")
    ws.cell(row=r, column=5, value=f"=SUM(E7:E{r-1})")
    ws.cell(row=r, column=6, value=f"=SUM(F7:F{r-1})")
    for c in (2, 3, 4, 5, 6):
        ws.cell(row=r, column=c).number_format = money_fmt if c != 2 else count_fmt
    total_row = r

    # --- Section 2: By service type ---
    start = total_row + 3
    ws.cell(row=start, column=1, value="2. Chi tiết theo loại dịch vụ").font = title_font
    svc_headers = ["Loại dịch vụ", "Số ca", "Bảo lãnh đủ", "Bảo lãnh 1 phần", "Không bảo lãnh"]
    for col, h in enumerate(svc_headers, start=1):
        cell = ws.cell(row=start + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    sr = start + 2
    for service in SERVICES:
        ws.cell(row=sr, column=1, value=service).border = border
        ws.cell(row=sr, column=2, value=f'=COUNTIF(DuLieu!$C$4:$C${last_row},A{sr})').border = border
        ws.cell(row=sr, column=3, value=(
            f'=COUNTIFS(DuLieu!$C$4:$C${last_row},A{sr},'
            f'DuLieu!$I$4:$I${last_row},"Bảo lãnh đủ")'
        )).border = border
        ws.cell(row=sr, column=4, value=(
            f'=COUNTIFS(DuLieu!$C$4:$C${last_row},A{sr},'
            f'DuLieu!$I$4:$I${last_row},"Bảo lãnh một phần")'
        )).border = border
        ws.cell(row=sr, column=5, value=(
            f'=COUNTIFS(DuLieu!$C$4:$C${last_row},A{sr},'
            f'DuLieu!$I$4:$I${last_row},"Không bảo lãnh (KH đã TT)")+'
            f'COUNTIFS(DuLieu!$C$4:$C${last_row},A{sr},'
            f'DuLieu!$I$4:$I${last_row},"Không bảo lãnh (Chưa thu)")'
        )).border = border
        for c in range(2, 6):
            ws.cell(row=sr, column=c).number_format = count_fmt
        sr += 1

    ws.cell(row=sr, column=1, value="TỔNG").font = header_font
    for c in range(1, 6):
        ws.cell(row=sr, column=c).fill = total_fill
        ws.cell(row=sr, column=c).border = border
        ws.cell(row=sr, column=c).font = header_font
    ws.cell(row=sr, column=2, value=f"=SUM(B{start+2}:B{sr-1})")
    ws.cell(row=sr, column=3, value=f"=SUM(C{start+2}:C{sr-1})")
    ws.cell(row=sr, column=4, value=f"=SUM(D{start+2}:D{sr-1})")
    ws.cell(row=sr, column=5, value=f"=SUM(E{start+2}:E{sr-1})")

    ws.column_dimensions["A"].width = 30
    for col in "BCDEFG":
        ws.column_dimensions[col].width = 18

    ws["A" + str(sr + 2)] = (
        "Ghi chú: Cột 'KH công nợ' là số tiền chưa thu từ khách hàng, "
        "không tính vào 'KH đã thanh toán'."
    )
    ws["A" + str(sr + 2)].font = Font(italic=True, color="666666")


def build_formula_sheet(wb, last_row):
    ws = wb.create_sheet("CongThuc")
    ws["A1"] = "Hướng dẫn cột & công thức"
    ws["A1"].font = title_font

    lines = [
        ("", ""),
        ("Cột trên sheet DuLieu", "Mô tả"),
        ("A – STT", "Số thứ tự"),
        ("B – Biển kiểm soát", "Mã định danh ca cứu hộ (1 dòng = 1 ca)"),
        ("C – Loại dịch vụ", "Cẩu kéo / Hỗ trợ tại chỗ / …"),
        ("D – Trạng thái bảo lãnh (nguồn)", "Giá trị export từ hệ thống"),
        ("E – PTI thanh toán", "Số tiền PTI đã/ sẽ thanh toán cho ca này"),
        ("F – KH đã thanh toán", "Số tiền đã thu từ khách hàng"),
        ("G – KH công nợ", "Số tiền KH chưa thanh toán (nếu có)"),
        ("H – Tổng chi phí", "=E+F+G"),
        ("I – Nhóm báo cáo", "Chuẩn hóa tên nhóm để pivot/báo cáo"),
        ("", ""),
        ("Công thức cột H (Tổng chi phí)", "=E2+F2+G2"),
        ("Công thức cột I (Nhóm báo cáo)", (
            '=IF(D2="PTI bảo lãnh","Bảo lãnh đủ",'
            'IF(D2="PTI bảo lãnh 1 phần","Bảo lãnh một phần",'
            'IF(D2="PTI không bảo lãnh","Không bảo lãnh (KH đã TT)",'
            'IF(D2="PTI không bảo lãnh, Khách hàng chưa thanh toán",'
            '"Không bảo lãnh (Chưa thu)",D2))))'
        )),
        ("", ""),
        ("Tổng số ca (ví dụ)", f"=COUNTA(DuLieu!B4:B{last_row})"),
        ("Tổng PTI thanh toán", f"=SUM(DuLieu!E4:E{last_row})"),
        ("Tổng KH đã TT", f"=SUM(DuLieu!F4:F{last_row})"),
        ("Tổng công nợ KH", f"=SUM(DuLieu!G4:G{last_row})"),
        ("Số ca bảo lãnh đủ", f'=COUNTIF(DuLieu!I4:I{last_row},"Bảo lãnh đủ")'),
        ("", ""),
        ("Lưu ý", "Mỗi ca = 1 dòng. Paste dữ liệu từ hàng 4, kéo công thức H:I xuống."),
        ("", "Cập nhật LAST_ROW trên sheet DuLieu hoặc dùng Excel Table (Ctrl+T) để công thức tự mở rộng."),
    ]

    for i, (a, b) in enumerate(lines, start=1):
        ws.cell(row=i, column=1, value=a)
        ws.cell(row=i, column=2, value=b)
        if i == 2:
            ws.cell(row=i, column=1).font = header_font
            ws.cell(row=i, column=2).font = header_font

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 70


def main():
    wb = Workbook()
    last_row = build_data_sheet(wb)
    build_report_sheet(wb, last_row)
    build_formula_sheet(wb, last_row)
    wb.save(OUTPUT)
    print(f"Created: {OUTPUT} (last_row={last_row})")


if __name__ == "__main__":
    main()
